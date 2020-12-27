import networkx as nx
import json
import numpy as np
from math import pi, radians, cos, sin, asin, sqrt
from openpyxl import load_workbook
import math
import datetime
from openpyxl import Workbook

def haversine(lon1, lat1, lon2, lat2):
    # 经度1，纬度1，经度2，纬度2 （十进制度数）
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # 将十进制度数转化为弧度
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
 
    # haversine公式
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 
    r = 6371 # 地球平均半径，单位为公里
    return c * r * 1000

def shp2graph(path_shp):
    H = nx.read_shp(path_shp)
    list_all_path = []
    set_all_node = set()
    dict_node_name = {}
    dict_name_node = {}
    list_node_name = []
    for i in [i for i in H.edges.data()]:
        info = json.loads(i[2]['Json'])['coordinates']
        info = tuple((tuple(i) for i in info))
        list_all_path.append(info)
        for j in info:
            set_all_node.add(j)
    for i in range(len(set_all_node)):
        dict_node_name[list(set_all_node)[i]] = 'v%s'%(i+1)
        dict_name_node['v%s'%(i+1)] = list(set_all_node)[i]
        list_node_name.append('v%s'%(i+1))

    list_row = []
    list_col = []
    list_value = []
    for edge in list_all_path:
        for i in range(len(edge)-1):
            list_row.append(dict_node_name[edge[i]])
            list_col.append(dict_node_name[edge[i+1]])
            list_value.append(int(haversine(edge[i][0], edge[i][1], edge[i+1][0], edge[i+1][1])))

    # print(list_node_name)
    #定义节点数
    nodes=np.array(list_node_name)
    #定义节点间的距离
    row=np.array(list_row)
    col=np.array(list_col)
    value=np.array(list_value)
    #生成无向图
    G=nx.Graph()
    #给图添加节点
    for i in range(0,np.size(nodes)):
        G.add_node(nodes[i])
    #添加带权的边
    for i in range(0,np.size(row)):
        G.add_weighted_edges_from([(row[i],col[i],value[i])])

    list_connected_names = []
    list_connected_nodes = []
    for node in list_node_name:
        nodes = nx.shortest_path(G,node).keys()
        if len(nodes)==304:
            list_connected_names = sorted(list(nodes))
            # print(b)
            break
    list_connected_nodes = [dict_name_node[i] for i in list_connected_names]

    return G, dict_node_name, dict_name_node, list_connected_nodes

def find_nearest(input_tuple, list_connected_nodes):
    output_tuple = tuple()
    distance_min = 10000000000000
    for node in list_connected_nodes:
        distance = haversine(node[0], node[1], input_tuple[0], input_tuple[1])
        if distance<distance_min:
            distance_min = distance
            output_tuple = node
    return output_tuple

def gen_path(G, input_tuple_source, input_tuple_target, dict_node_name, dict_name_node, list_connected_nodes):
    start = dict_node_name[find_nearest(input_tuple_source, list_connected_nodes)]
    # print('从%s出发，位置是%s'%(start, dict_name_node[start]))
    end = dict_node_name[find_nearest(input_tuple_target, list_connected_nodes)]
    # print('到%s为止，位置是%s'%(end, dict_name_node[end]))
    path=nx.dijkstra_path(G, source=start, target=end)
    # print('出发点{0}到结束点{1}的路径：'.format(start,end), path)
    # print('出发点{0}到结束点{1}的路径（用坐标表示）：'.format(start,end), [dict_name_node[i] for i in path])
    distance=nx.dijkstra_path_length(G, source=start, target=end)
    # print('出发点{0}到结束点{1}的距离约为{2}米'.format(start,end,distance))
    path_loc = [dict_name_node[i] for i in path]
    return start, end, path, path_loc, distance

def xlsx2poi(path_xlsx):
    poi_out = {}
    wb = load_workbook(path_xlsx, read_only=True)
    name_list = wb.get_sheet_names()
    for name in name_list:
        my_sheet = wb.get_sheet_by_name(name)
        len_sheet = my_sheet.max_row
        tag = my_sheet.title.split('.')[1]
        poi_out[tag] = []
        # my_sheet = wb.active
        for i in range(len_sheet-1):
            if my_sheet['B%s'%(i+2)].value!='' and my_sheet['C%s'%(i+2)].value!='':
                # poi_out.append((my_sheet['B%s'%(i+2)].value, my_sheet['C%s'%(i+2)].value, tag))
                poi_out[tag].append((my_sheet['B%s'%(i+2)].value, my_sheet['C%s'%(i+2)].value))
            else:
                break
    return poi_out

def calc_azimuth(lat1, lon1, lat2, lon2):
    lat1_rad = lat1 * math.pi / 180
    lon1_rad = lon1 * math.pi / 180
    lat2_rad = lat2 * math.pi / 180
    lon2_rad = lon2 * math.pi / 180

    y = math.sin(lon2_rad - lon1_rad) * math.cos(lat2_rad)
    x = math.cos(lat1_rad) * math.sin(lat2_rad) - \
        math.sin(lat1_rad) * math.cos(lat2_rad) * math.cos(lon2_rad - lon1_rad)

    brng = math.atan2(y, x) * 180 / math.pi
    degree_out = float((brng + 360.0) % 360.0)
    str_out = []

    if degree_out<90:
        str_out = ['北偏东方向', degree_out]
    elif degree_out<180:
        str_out = ['东偏南方向', degree_out-90]
    elif degree_out<270:
        str_out = ['南偏西方向', degree_out-180]
    else:
        str_out = ['西偏北方向', degree_out-270]

    return str_out

list_poi = xlsx2poi('./programmingdata/poi数据集.xlsx')
# print(list_poi)

G, dict_node_name, dict_name_node, list_connected_nodes = shp2graph('./programmingdata/路网/全路网.shp')

# start, end, path, path_loc, distance = gen_path(G, (121.481575662451914, 31.023043916598567), (121.421659419479283, 31.922859270424533), dict_node_name, dict_name_node, list_connected_nodes)

mode = input('请选择展示两地点间路线或从当前位置出发，展示两地间路线请输入1，从当前位置出发请输入2:\n')
if mode=='1':
    start_name = input('可供选择的起始地点有：%s，请输入您的起始地点：\n'%(list(list_poi.keys())))
    end_name = input('可供选择的终止地点有：%s，请输入您的终止地点：\n'%(list(list_poi.keys())))
    print('正在检索路线哟！')
    start_loc = list_poi[start_name][0]
    end_loc = list_poi[end_name][0]
    start, end, path, path_loc, distance = gen_path(G, start_loc, end_loc, dict_node_name, dict_name_node, list_connected_nodes)
    print('找到辣！走完全程需要%s米呢！'%(distance))
    print('完整的路线按标记展示为：\n%s'%(path))
    mode_by_loc = input('若需要按经纬度坐标展示，请输入1，否则请回车或输入任意字符\n')
    if mode_by_loc=='1':
        print('精确坐标路径来辣！请查收：\n%s'%(path_loc))
        print('拜拜～')
    else:
        print('拜拜～')
elif mode=='2':
    start_loc = input('请输入您的当前坐标（举个栗子：%s）\n'%('(121.42229714275993, 31.023261634350888)'))
    start_loc = tuple(json.loads('['+start_loc[1:-1]+']'))
    end_name = input('可供选择的终止地点有：%s，请输入您的终止地点：\n'%(list(list_poi.keys())))
    end_loc = find_nearest(start_loc, list_poi[end_name])
    start, end, path, path_loc, distance = gen_path(G, start_loc, end_loc, dict_node_name, dict_name_node, list_connected_nodes)
    mode_show = input('我们为您找到了最近的%s，若直接展示整条路线，请输入1，若逐步展示，请输入2:\n'%(end_name))
    if mode_show=='1':
        print('找到辣！走完全程需要%s米呢！'%(distance))
        print('完整的路线按标记展示为：\n%s'%(path))
        mode_by_loc = input('若需要按经纬度坐标展示，请输入1，否则请回车或输入任意字符\n')
        if mode_by_loc=='1':
            print('精确坐标路径来辣！请查收：\n%s'%(path_loc))
            print('拜拜～')
        else:
            print('拜拜～')
    elif mode_show=='2':
        print('找到辣！走完全程需要%s米呢！'%(distance))
        print('出发惹！')
        for i in range(len(path_loc)-1):
            lat1, lon1, lat2, lon2 = path_loc[i][0],  path_loc[i][1],  path_loc[i+1][0],  path_loc[i+1][1]
            face_degree = calc_azimuth(lat1, lon1, lat2, lon2)
            print('您需要向 %s %s 度前进%s米'%(face_degree[0], face_degree[1], haversine(lat1, lon1, lat2, lon2)))
            feedback = input('走到了告诉我一下哟，回车就可以辣')
        print('到啦～拜拜')

gen_xlsx = input('是否需要生成Excel文件方便绘图呢？需要请输入1，不需要回车就好\n')
if gen_xlsx=='1':
    path_xlsx = './result_xlsx/%s.xlsx'%(str(datetime.datetime.now()).replace(' ','-').replace(':','-')[:16])

    list_write = path_loc

    wb = Workbook()
    sheet = wb.active
    sheet['A%s'%(1)] = '经度'
    sheet['B%s'%(1)] = '纬度'
    for i in range(len(list_write)):
        sheet['A%s'%(i+2)] = list_write[i][0]
        sheet['B%s'%(i+2)] = list_write[i][1]
    wb.save(path_xlsx)
    print('生成啦，它躺在result_xlsx文件夹里哦，按照时间命名惹')